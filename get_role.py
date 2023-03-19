import discord
import random
import openpyxl
import gspread
from google.oauth2.service_account import Credentials
from discord.ext import tasks
from discord.ext import commands
from discord import Option


# Discord API認証トークンを設定する
TOKEN = '***'
SERVER_ID = ***

intents = discord.Intents.default()
intents = discord.Intents.all()
#client = discord.Client(intents=intents)

bot = discord.Bot(intents=intents)

# お決まりの文句
# 2つのAPIを記述しないとリフレッシュトークンを3600秒毎に発行し続けなければならない
scope = ['https://www.googleapis.com/auth/spreadsheets','https://www.googleapis.com/auth/drive']
#ダウンロードしたjsonファイル名をクレデンシャル変数に設定。
credentials = Credentials.from_service_account_file("get-roles-631c3dc0a756.json", scopes=scope)
#OAuth2の資格情報を使用してGoogle APIにログイン。
gc = gspread.authorize(credentials)
#スプレッドシートIDを変数に格納する。
SPREADSHEET_KEY = '11JBh5AQwjk0Ktk3yaq8O9X10_nRgVTESMQ3BMB1-1Uc'
# スプレッドシート(ブック)を開く
workbook = gc.open_by_key(SPREADSHEET_KEY)

random_contents = [
    "にゃーん",
    "わん！",
    "コケッコッコー",
]

cnt_reaction = {'hoge':0}

# サーバーのメンバーリストを取得する関数
async def get_members(guild):
    members = []
    for member in guild.members:
        roles = [role.name for role in member.roles]# if role.name != "@everyone"]
        name = member.name + '#' + member.discriminator
        for i, role in enumerate(roles):
            members.append([name, role])

    return members

def out_roles(members):
    # スプレッドシート(ブック)を開く
    workbook = gc.open_by_key(SPREADSHEET_KEY)
    # シートの一覧を取得する。(リスト形式)
    worksheets = workbook.worksheets()
    print(worksheets)
    # シートを開く
    worksheet = workbook.worksheet('role_list')
    workbook.values_clear("!A1:H100")

    worksheet.update_cell(1, 1, 'Username')
    worksheet.update_cell(1, 2, 'Roles')

    # メンバーリストを表示する
    cnt = 2
    for i, member in enumerate(members):
        worksheet.update_cell(i+2, 1, member[0])
        worksheet.update_cell(i+2, 2, member[1])
        print(i)
        print(member)

@bot.event
async def on_message(message):
    # 送信者がbotである場合は弾く
    if message.author.bot:
        return 
    # メッセージの本文が 鳴いて だった場合
    if message.content == "鳴いて":
        # 送信するメッセージをランダムで決める
        content = random.choice(random_contents)
        # メッセージが送られてきたチャンネルに送る
        await message.channel.send(content)
    elif message.content == "おはよう":
        await message.channel.send("おはよう！！")

@bot.event
async def on_reaction_add(reaction, user):
    global cnt_reaction
    channel_name = str(reaction.message.channel.name)
    reaction_name = str(reaction.emoji)

    #cnt_reaction[channel_id] += 1
    
    # スプレッドシート(ブック)を開く
    workbook = gc.open_by_key(SPREADSHEET_KEY)
    # シートの一覧を取得する。(リスト形式)
    worksheets = workbook.worksheets()
    worksheet = workbook.worksheet('reaction_cnt')
    
    print(worksheets)

    flg = False
    # 1行目からユーザ名を検索する
    for i in range(1, worksheet.row_count):
        
        if flg == True:
            break

        print(i)
        if worksheet.cell(i, 1).value == channel_name:
            # 既存のユーザーが見つかった場合、カウントをインクリメントする
            for j in range(1, worksheet.col_count):
                if worksheet.cell(1, j).value == reaction_name:
                    if worksheet.cell(i, j).value is None:
                        count = 1
                    else:
                        count = int(worksheet.cell(i, j).value) + 1
                    
                    worksheet.update_cell(i, j, count)
                    flg = True
                    break

                elif worksheet.cell(1, j).value is None:
                    worksheet.update_cell(1, j, reaction_name)
                    worksheet.update_cell(i, j, 1)
                    flg = True
                    break
                                    
        elif worksheet.cell(i, 1).value is None:
            worksheet.update_cell(i, 1, channel_name)
            
            for j in range(1, worksheet.col_count):
                if worksheet.cell(1, j).value == reaction_name:
                    worksheet.update_cell(i, j, 1)
                    flg = True
                    break

                elif worksheet.cell(1, j).value is None:
                    worksheet.update_cell(1, j, reaction_name)
                    worksheet.update_cell(i, j, 1)
                    flg = True
                    break

    print("add")

@bot.event
async def on_reaction_remove(reaction, user):
    
    global cnt_reaction
    channel_id = str(reaction.message.channel.id)
    channel_name = str(reaction.message.channel.name)
    reaction_name = str(reaction.emoji)
    #cnt_reaction[channel_id] -= 1
    
    # スプレッドシート(ブック)を開く
    workbook = gc.open_by_key(SPREADSHEET_KEY)
    # シートの一覧を取得する。(リスト形式)
    worksheets = workbook.worksheets()
    worksheet = workbook.worksheet('reaction_cnt')
    print(worksheets)
    # シートを開く


    flg = False
    
    # 1行目からユーザ名を検索する
    for i in range(1, worksheet.row_count):
        
        if flg == True:
            break

        print(i)
        if worksheet.cell(i, 1).value == channel_name:
            # 既存のユーザーが見つかった場合、カウントをインクリメントする
            for j in range(1, worksheet.col_count):
                if worksheet.cell(1, j).value == reaction_name:
                    if worksheet.cell(i, j).value is None:
                        count = 0
                    else:
                        count = int(worksheet.cell(i, j).value) - 1

                    worksheet.update_cell(i, j, count)
                    flg = True
                    break

                elif worksheet.cell(1, j).value is None:
                    worksheet.update_cell(1, j, reaction_name)
                    flg = True
                    break
                                    
        elif worksheet.cell(i, 1).value is None:
            worksheet.update_cell(i, 1, channel_name)
            
            for j in range(1, worksheet.col_count):
                if worksheet.cell(1, j).value == reaction_name:
                    worksheet.update_cell(i, j, 1)
                    flg = True
                    break

                elif worksheet.cell(1, j).value is None:
                    worksheet.update_cell(1, j, reaction_name)
                    flg = True
                    break

@bot.event
async def on_ready():
    print(f"{bot.user} コマンド待機中...")

@bot.slash_command(description="こんにちわコマンド", guild_ids=[SERVER_ID])
async def hello(
    ctx: discord.ApplicationContext,
    name: Option(str, required=False, description="input name")
):
    name = name or ctx.author.name
    await ctx.respond(f"{name} さん, こんにちは。")
 
@bot.slash_command(description="こんばんわコマンド", guild_ids=[SERVER_ID])
async def goodnight(
    ctx: discord.ApplicationContext,
    name: Option(str, required=False, description="input name")
):
    name = name or ctx.author.name
    await ctx.respond(f"{name} さん, こんばんわ。")

@bot.slash_command(description="ロール取得コマンド", guild_ids=[SERVER_ID])
async def get_role(ctx):
    
    # スプレッドシート(ブック)を開く
    workbook = gc.open_by_key(SPREADSHEET_KEY)
    # シートを開く
    worksheet = workbook.worksheet('role_list')
    worksheet.clear()

    worksheet.update_cell(1, 1, 'Username')
    worksheet.update_cell(1, 2, 'Roles')

    members = await get_members(ctx.guild)

    cnt = 0
    for member in members:
        if member[1] != "@everyone":
            print(member)
            worksheet.update_cell(cnt+2, 1, member[0])
            worksheet.update_cell(cnt+2, 2, member[1])
            cnt += 1

    print("hogehoge")
    await ctx.send("メンバー出力完了:https://docs.google.com/spreadsheets/d/11JBh5AQwjk0Ktk3yaq8O9X10_nRgVTESMQ3BMB1-1Uc/edit#gid=0")

@bot.slash_command(description="リアクション数取得コマンド", guild_ids=[SERVER_ID])
async def get_reac_cnt(ctx):

    await ctx.send("リアクション数取得完了:https://docs.google.com/spreadsheets/d/11JBh5AQwjk0Ktk3yaq8O9X10_nRgVTESMQ3BMB1-1Uc/edit#gid=91442236")


#client.run(TOKEN)
bot.run(TOKEN)
